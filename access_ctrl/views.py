from django.db import IntegrityError
from django.db.models import Q, Max, Count
from django.db.models.functions import TruncDay
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.generics import ListAPIView, UpdateAPIView
from rest_framework.permissions import IsAuthenticated
from datetime import timedelta, datetime, time
from .models import Visita, Acceso, ProhibicionAcceso, Acceso
from core.models import Instalacion, Sector, Empresa
from core.serializers import SectorSer
from .serializers import AccesoSerializer
from .serializers import IngresoRequest, SalidaRequest, AccesoSerializer, VisitaSerializer, VisitaSimpleSerializer, \
    AccesoFullSerializer, EnrolamientoSerializer, CargaMasivaEnrolamientoSerializer
from drf_spectacular.utils import extend_schema
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment


def es_admin_general(user):
    return bool(user.empresa and user.empresa.es_administradora_general)


class AccesoListView(ListAPIView):
    serializer_class = AccesoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = Acceso.objects.select_related(
            "visita",
            "instalacion",
            "sector",
            "empresa",
            "guardia",
        ).all()

        visita_id = self.request.query_params.get("visita_id")
        instalacion_id = self.request.query_params.get("instalacion_id")
        empresa_id = self.request.query_params.get("empresa_id")
        tipo = self.request.query_params.get("tipo")

        if es_admin_general(user):
            if empresa_id:
                queryset = queryset.filter(empresa_id=empresa_id)
            if instalacion_id:
                queryset = queryset.filter(instalacion_id=instalacion_id)

        elif user.role == "admin":
            queryset = queryset.filter(empresa_id=user.empresa_id)
            if instalacion_id:
                queryset = queryset.filter(instalacion_id=instalacion_id)

        elif user.role == "guardia":
            queryset = queryset.filter(
                empresa_id=user.empresa_id,
                instalacion_id=user.instalacion_id
            )

        else:
            return Acceso.objects.none()

        if visita_id:
            queryset = queryset.filter(visita_id=visita_id)

        if tipo in ["ingreso", "salida"]:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by("-fecha_hora")


def _get_visita(payload):
    # 1) por id
    if payload.get("visita_id"):
        return Visita.objects.filter(id=payload["visita_id"]).first()

    # 2) por documento
    if payload.get("es_extranjero"):
        dni = (payload.get("dni_extranjero") or "").strip()
        if dni:
            return Visita.objects.filter(dni_extranjero=dni, es_extranjero=True).first()
    else:
        rut = (payload.get("rut") or "").strip()
        if rut:
            return Visita.objects.filter(rut=rut, es_extranjero=False).first()
    return None


def _crear_o_actualizar_visita(payload):
    v = _get_visita(payload)
    if v:
        # completar datos faltantes si vienen
        for f in ["nombre", "apellido", "empresa", "patente"]:
            val = payload.get(f)
            if val and not getattr(v, f):
                setattr(v, f, val)
        v.save()
        return v, False
    # crear
    v = Visita.objects.create(
        rut=payload.get("rut") if not payload.get("es_extranjero") else None,
        dni_extranjero=payload.get("dni_extranjero") if payload.get("es_extranjero") else None,
        es_extranjero=payload.get("es_extranjero") or False,
        nombre=payload.get("nombre") or "Sin nombre",
        apellido=payload.get("apellido") or "",
        empresa=payload.get("empresa") or "",
        patente=payload.get("patente") or "",
    )
    return v, True


def _hay_prohibicion(v, instalacion):
    now = timezone.now()
    return ProhibicionAcceso.objects.filter(
        visita=v, instalacion=instalacion
    ).filter(Q(fecha_fin__isnull=True, fecha_inicio__lte=now) | Q(fecha_inicio__lte=now, fecha_fin__gte=now)).exists()


def _ultimo_evento(v, instalacion):
    return Acceso.objects.filter(visita=v, instalacion=instalacion).order_by("-fecha_hora").first()


class IngresoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(request=IngresoRequest, responses={201: AccesoSerializer})
    def post(self, request):
        ser = IngresoRequest(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # ✅ 1️⃣ Obtener instalación desde el usuario logueado
        user = request.user
        if not user.instalacion:
            return Response(
                {"ok": False, "error": "usuario_sin_instalacion_asociada"},
                status=status.HTTP_400_BAD_REQUEST
            )

        instalacion = user.instalacion

        # ✅ 2️⃣ Obtener el sector por ID
        sector_id = data.get("sector_id")
        if not sector_id:
            return Response(
                {"ok": False, "error": "sector_id_requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            sector = Sector.objects.get(id=sector_id, instalacion=instalacion)
        except Sector.DoesNotExist:
            return Response(
                {"ok": False, "error": "sector_no_valido"},
                status=status.HTTP_404_NOT_FOUND
            )

        # ✅ 3️⃣ Crear o actualizar visita
        visita, created = _crear_o_actualizar_visita(data)

        # ✅ 4️⃣ Verificar prohibición
        if _hay_prohibicion(visita, instalacion):
            return Response(
                {"ok": False, "error": "prohibido"},
                status=status.HTTP_403_FORBIDDEN
            )

        # ✅ 5️⃣ Evitar doble ingreso
        last = _ultimo_evento(visita, instalacion)
        if last and last.tipo == "ingreso":
            return Response(
                {"ok": False, "error": "visita_ya_adentro"},
                status=status.HTTP_409_CONFLICT
            )

        # ✅ 6️⃣ Registrar acceso
        acceso = Acceso.objects.create(
            visita=visita,
            instalacion=instalacion,
            sector=sector,
            tipo="ingreso",
            fecha_hora=timezone.now(),
            comentario=data.get("comentario") or "",
            guardia=user,
            empresa=instalacion.empresa,
        )

        return Response(
            {"ok": True, "mensaje": "Ingreso registrado", "acceso": AccesoSerializer(acceso).data},
            status=201
        )


class SalidaView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(request=SalidaRequest, responses={201: AccesoSerializer})
    def post(self, request):
        ser = SalidaRequest(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # ✅ Aquí sí puedes acceder al usuario
        user = request.user
        instalacion = user.instalacion
        if not instalacion:
            return Response(
                {"ok": False, "error": "usuario_sin_instalacion_asociada"},
                status=400
            )

        sector_id = data.get("sector_id")
        if not sector_id:
            return Response(
                {"ok": False, "error": "sector_id_requerido"},
                status=400
            )

        try:
            sector = Sector.objects.get(id=sector_id, instalacion=instalacion)
        except Sector.DoesNotExist:
            return Response(
                {"ok": False, "error": "sector_no_valido"},
                status=404
            )

        # Aquí continúas con el flujo normal:
        visita = _get_visita(data)
        if not visita:
            return Response({"ok": False, "error": "visita_no_encontrada"}, status=404)

        last = _ultimo_evento(visita, instalacion)
        if not last or last.tipo != "ingreso":
            return Response({"ok": False, "error": "no_hay_ingreso_abierto"}, status=409)

        acceso = Acceso.objects.create(
            visita=visita,
            instalacion=instalacion,
            sector=sector,
            tipo="salida",
            fecha_hora=timezone.now(),
            comentario=data.get("comentario") or "",
            foto_url=data.get("foto_url") or "",
            guardia=user,
            empresa=instalacion.empresa,
        )

        return Response(
            {"ok": True, "mensaje": "Salida registrada", "acceso": AccesoSerializer(acceso).data},
            status=201
        )


class BuscarPorRUTView(APIView):
    def get(self, request, rut):
        try:
            visita = Visita.objects.get(rut=rut)
            prohibicion = ProhibicionAcceso.objects.filter(visita=visita).exists()
            if prohibicion:
                return Response({
                    "ok": False,
                    "mensaje": "Acceso prohibido",
                    "visita": VisitaSerializer(visita).data
                }, status=status.HTTP_403_FORBIDDEN)
            return Response({
                "ok": True,
                "mensaje": "Visita encontrada",
                "visita": VisitaSerializer(visita).data
            }, status=status.HTTP_200_OK)
        except Visita.DoesNotExist:
            return Response({
                "ok": False,
                "mensaje": "No se encontró un visitante con ese RUT"
            }, status=status.HTTP_404_NOT_FOUND)


class BuscarPorDNIView(APIView):
    def get(self, request, dni):
        try:
            visita = Visita.objects.get(dni_extranjero=dni)
            prohibicion = ProhibicionAcceso.objects.filter(visita=visita).exists()
            if prohibicion:
                return Response({
                    "ok": False,
                    "mensaje": "Acceso prohibido",
                    "visita": VisitaSerializer(visita).data
                }, status=status.HTTP_403_FORBIDDEN)
            return Response({
                "ok": True,
                "mensaje": "Visita encontrada",
                "visita": VisitaSerializer(visita).data
            }, status=status.HTTP_200_OK)
        except Visita.DoesNotExist:
            return Response({
                "ok": False,
                "mensaje": "No se encontró un visitante con ese DNI"
            }, status=status.HTTP_404_NOT_FOUND)


class RegistrarVisitaView(APIView):
    """
    Crea una nueva visita o actualiza datos mínimos si ya existe.
    """
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(request=VisitaSerializer, responses={201: VisitaSerializer})
    def post(self, request):
        data = request.data

        # Determinar si es extranjero o no
        es_extranjero = bool(data.get("dni_extranjero"))
        data["es_extranjero"] = es_extranjero

        # Crear o actualizar usando tu función auxiliar
        visita, creada = _crear_o_actualizar_visita(data)

        serializer = VisitaSerializer(visita)
        mensaje = "Visita creada correctamente" if creada else "Visita actualizada correctamente"

        return Response(
            {"ok": True, "mensaje": mensaje, "visita": serializer.data},
            status=status.HTTP_201_CREATED if creada else status.HTTP_200_OK
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def buscar_ultimo_acceso_por_rut(request, rut):
    """
    Devuelve el último registro de acceso (ingreso o salida) asociado al RUT.
    Si no hay ingreso previo abierto, informa que no se puede registrar salida.
    Incluye información del sector visitado y si requiere documentación de salida.
    """
    try:
        visita = Visita.objects.get(rut=rut)
    except Visita.DoesNotExist:
        return Response(
            {"ok": False, "mensaje": "No existe una visita registrada con ese RUT."},
            status=status.HTTP_404_NOT_FOUND
        )

    # Buscar último acceso registrado
    ultimo = Acceso.objects.filter(visita=visita).order_by("-fecha_hora").first()

    if not ultimo:
        return Response(
            {"ok": False, "mensaje": "No hay registros de accesos para esta visita."},
            status=status.HTTP_404_NOT_FOUND
        )

    # ⚙️ Obtener información del sector y si requiere documentación
    requiere_doc = ultimo.sector.requiere_guia if hasattr(ultimo.sector, "requiere_guia") else False
    sector_info = {
        "id": ultimo.sector.id,
        "nombre": ultimo.sector.nombre,
        "requiere_documentacion": requiere_doc
    }

    # 🚫 Si el último evento fue una salida → no permitir otra salida
    if ultimo.tipo == "salida":
        return Response(
            {
                "ok": False,
                "mensaje": "La visita no tiene un ingreso abierto.",
                "ultimo_acceso": AccesoSerializer(ultimo).data,
                "sector": sector_info
            },
            status=status.HTTP_409_CONFLICT
        )

    # ✅ Si el último evento fue un ingreso → permitir registrar salida
    return Response(
        {
            "ok": True,
            "mensaje": "Ingreso encontrado. Puede registrar salida.",
            "ultimo_acceso": AccesoSerializer(ultimo).data,
            "sector": sector_info
        },
        status=status.HTTP_200_OK
    )


class AccesosUltimas24View(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AccesoSerializer

    def get_queryset(self):
        user = self.request.user
        now = timezone.localtime()
        start = now - timedelta(hours=24)

        qs = Acceso.objects.select_related(
            "visita",
            "instalacion",
            "sector",
            "empresa",
            "guardia"
        ).filter(
            fecha_hora__gte=start
        )

        empresa_id = self.request.query_params.get("empresa_id")
        instalacion_id = self.request.query_params.get("instalacion_id")

        if es_admin_general(user):
            if empresa_id:
                qs = qs.filter(empresa_id=empresa_id)
            if instalacion_id:
                qs = qs.filter(instalacion_id=instalacion_id)

        elif user.role == "admin":
            qs = qs.filter(empresa_id=user.empresa_id)
            if instalacion_id:
                qs = qs.filter(instalacion_id=instalacion_id)

        elif user.role == "guardia":
            qs = qs.filter(
                empresa_id=user.empresa_id,
                instalacion_id=user.instalacion_id
            )

        else:
            qs = qs.none()

        return qs.order_by("-fecha_hora")

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        total = queryset.count()
        total_ingresos = queryset.filter(tipo="ingreso").count()
        total_salidas = queryset.filter(tipo="salida").count()

        return Response({
            "ok": True,
            "total": total,
            "total_ingresos": total_ingresos,
            "total_salidas": total_salidas,
            "results": serializer.data
        })


class AccesosDiaEnCursoView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AccesoSerializer

    def get_queryset(self):
        user = self.request.user
        now = timezone.localtime()

        start = timezone.make_aware(
            datetime.combine(now.date(), time(6, 0)),
            timezone.get_current_timezone()
        )
        next_midnight = timezone.make_aware(
            datetime.combine(now.date() + timedelta(days=1), time(0, 0)),
            timezone.get_current_timezone()
        )

        qs = Acceso.objects.select_related(
            "visita", "instalacion", "sector", "empresa", "guardia"
        ).filter(
            fecha_hora__gte=start,
            fecha_hora__lt=next_midnight
        )

        empresa_id = self.request.query_params.get("empresa_id")
        instalacion_id = self.request.query_params.get("instalacion_id")

        if es_admin_general(user):
            if empresa_id:
                qs = qs.filter(empresa_id=empresa_id)
            if instalacion_id:
                qs = qs.filter(instalacion_id=instalacion_id)

        elif user.role == "admin":
            qs = qs.filter(empresa_id=user.empresa_id)
            if instalacion_id:
                qs = qs.filter(instalacion_id=instalacion_id)

        elif user.role == "guardia":
            qs = qs.filter(
                empresa_id=user.empresa_id,
                instalacion_id=user.instalacion_id
            )

        else:
            qs = qs.none()

        return qs.order_by("-fecha_hora")

class SectoresPorInstalacionView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = SectorSer

    def get_queryset(self):
        user = self.request.user
        inst_id = self.kwargs.get("instalacion_id")

        qs = Sector.objects.filter(instalacion_id=inst_id)

        if es_admin_general(user):
            return qs.order_by("nombre")

        return qs.filter(instalacion__empresa_id=user.empresa_id).order_by("nombre")


class VisitasPorInstalacionView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VisitaSerializer

    def get_queryset(self):
        user = self.request.user
        instalacion_id = self.kwargs.get("instalacion_id")

        qs = Visita.objects.filter(instalacion_id=instalacion_id)

        if not es_admin_general(user):
            qs = qs.filter(instalacion__empresa_id=user.empresa_id)

        q = self.request.query_params.get("q")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(apellido__icontains=q) |
                Q(rut__icontains=q) |
                Q(dni_extranjero__icontains=q)
            )

        return qs.order_by("-creado_en")


class AccesosPorMesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        year = int(request.query_params.get("year", timezone.localtime().year))
        month = int(request.query_params.get("month", timezone.localtime().month))

        empresa_id = request.query_params.get("empresa_id")
        instalacion_id = request.query_params.get("instalacion_id")

        base = Acceso.objects.filter(
            fecha_hora__year=year,
            fecha_hora__month=month,
        )

        if es_admin_general(user):
            if empresa_id:
                base = base.filter(empresa_id=empresa_id)
            if instalacion_id:
                base = base.filter(instalacion_id=instalacion_id)
        else:
            base = base.filter(empresa_id=user.empresa_id)
            if instalacion_id:
                base = base.filter(instalacion_id=instalacion_id)

        diario = (
            base.annotate(dia=TruncDay("fecha_hora"))
            .values("dia", "tipo")
            .annotate(total=Count("id"))
            .order_by("dia", "tipo")
        )

        include_detail = request.query_params.get("detail") == "1"
        data = {
            "year": year,
            "month": month,
            "empresa_id": empresa_id if es_admin_general(user) else user.empresa_id,
            "instalacion_id": instalacion_id,
            "resumen_diario": list(diario),
        }

        if include_detail:
            data["accesos"] = AccesoSerializer(
                base.select_related("visita", "sector", "instalacion", "empresa").order_by("-fecha_hora"),
                many=True
            ).data

        return Response({"ok": True, "data": data}, status=200)


class VisitaUpdateView(UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VisitaSerializer
    queryset = Visita.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user

        if es_admin_general(user):
            return qs

        return qs.filter(instalacion__empresa_id=user.empresa_id)


class AccesoUpdateAdminView(UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AccesoFullSerializer
    queryset = Acceso.objects.all()

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user

        if es_admin_general(user):
            return qs

        if user.is_admin():
            return qs.filter(empresa_id=user.empresa_id)

        return qs.none()

    def patch(self, request, *args, **kwargs):
        user = request.user

        if not user.is_admin() and not es_admin_general(user):
            return Response(
                {"ok": False, "error": "No tiene permisos para editar accesos"},
                status=status.HTTP_403_FORBIDDEN
            )

        return self.partial_update(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        user = request.user

        if not user.is_admin() and not es_admin_general(user):
            return Response(
                {"ok": False, "error": "No tiene permisos para editar accesos"},
                status=status.HTTP_403_FORBIDDEN
            )

        return self.update(request, *args, **kwargs)


class CargaMasivaAccesosView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        if not isinstance(data, list):
            return Response({"error": "Debe enviar una lista de accesos"}, status=400)

        creados = 0
        errores = []
        user = request.user

        for acceso_data in data:
            try:
                rut = acceso_data.get("rut")
                nombre = acceso_data.get("nombre") or "Sin nombre"

                visita, creada = Visita.objects.get_or_create(
                    rut=rut,
                    defaults={
                        "dni_extranjero": acceso_data.get("dni_extranjero", ""),
                        "es_extranjero": acceso_data.get("es_extranjero", False),
                        "nombre": nombre,
                        "apellido": acceso_data.get("apellido", ""),
                        "empresa": acceso_data.get("empresa", ""),
                        "patente": acceso_data.get("patente", ""),
                    },
                )

                if not visita.nombre:
                    visita.nombre = nombre
                    visita.save()

                serializer = AccesoSerializer(data={
                    "visita": visita.id,
                    "instalacion": acceso_data.get("instalacion_id"),
                    "sector": acceso_data.get("sector_id"),
                    "tipo": "ingreso",
                    "fecha_hora": timezone.now(),
                    "comentario": acceso_data.get("comentario", ""),
                    "empresa": user.empresa_id,
                    "guardia": user.id
                })

                if serializer.is_valid():
                    serializer.save()
                    creados += 1
                else:
                    errores.append(serializer.errors)

            except Exception as e:
                errores.append(str(e))

        return Response(
            {"ok": True, "total_creados": creados, "errores": errores[:10]},
            status=status.HTTP_201_CREATED
        )


class SectoresDisponiblesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # 🔥 usuario sectorial → solo su sector
        if user.solo_enrolamiento:
            sectores = Sector.objects.filter(id=user.sector_id)

        # 🔥 admin → sectores de su instalación
        elif user.instalacion_id:
            sectores = Sector.objects.filter(instalacion_id=user.instalacion_id)

        else:
            # superadmin
            sectores = Sector.objects.all()

        data = [
            {
                "id": s.id,
                "nombre": s.nombre
            }
            for s in sectores
        ]

        return Response(data)


class EnroladosListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # 🔥 cliente_sector → solo su sector
        if user.solo_enrolamiento:
            visitas = Visita.objects.filter(sector=user.sector)

        # 🔥 admin → por instalación
        elif user.instalacion_id:
            visitas = Visita.objects.filter(instalacion_id=user.instalacion_id)

        else:
            visitas = Visita.objects.all()

        serializer = EnrolamientoSerializer(visitas, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = EnrolamientoSerializer(
            data=request.data,
            context={"request": request}
        )

        serializer.is_valid(raise_exception=True)
        visita = serializer.save()

        return Response(
            EnrolamientoSerializer(visita).data,
            status=201
        )


class CargaMasivaEnrolamientoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = CargaMasivaEnrolamientoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        archivo = serializer.validated_data["archivo"]

        if not archivo.name.endswith(".xlsx"):
            return Response(
                {"detail": "Solo se permiten archivos .xlsx"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if user.solo_enrolamiento:
            sector = user.sector
            instalacion = user.instalacion
        else:
            sector_id = serializer.validated_data.get("sector_id")
            if not sector_id:
                return Response(
                    {"detail": "Debe enviar sector_id"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                sector = Sector.objects.get(id=sector_id)
            except Sector.DoesNotExist:
                return Response(
                    {"detail": "El sector enviado no existe"},
                    status=status.HTTP_404_NOT_FOUND
                )

            instalacion = sector.instalacion

        try:
            wb = load_workbook(filename=archivo)
            ws = wb.active
        except Exception:
            return Response(
                {"detail": "No se pudo leer el archivo Excel"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Buscar encabezados dinámicamente
        header_row_idx = None
        header_map = {}

        expected_headers = {
            "tipo documento",
            "rut",
            "dni",
            "nombre",
            "apellido",
            "patente",
            "comentario",
        }

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = []
            for cell in row:
                if cell is None:
                    normalized.append("")
                else:
                    normalized.append(str(cell).strip().lower().replace("_", " "))

            row_headers = set(x for x in normalized if x)

            if expected_headers.issubset(row_headers):
                header_row_idx = row_idx
                for col_idx, value in enumerate(normalized):
                    if value in expected_headers:
                        header_map[value] = col_idx
                break

        if not header_row_idx:
            return Response(
                {
                    "detail": "No se encontraron los encabezados requeridos.",
                    "encabezados_requeridos": [
                        "TIPO DOCUMENTO",
                        "RUT",
                        "DNI",
                        "NOMBRE",
                        "APELLIDO",
                        "PATENTE",
                        "COMENTARIO",
                    ]
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        total = 0
        creados = 0
        errores = []

        for idx, row in enumerate(
                ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1
        ):
            tipo_documento = str(row[header_map["tipo documento"]]).strip().upper() if row[header_map[
                "tipo documento"]] is not None else ""
            rut = str(row[header_map["rut"]]).strip() if row[header_map["rut"]] is not None else ""
            dni_extranjero = str(row[header_map["dni"]]).strip() if row[header_map["dni"]] is not None else ""
            nombre = str(row[header_map["nombre"]]).strip() if row[header_map["nombre"]] is not None else ""
            apellido = str(row[header_map["apellido"]]).strip() if row[header_map["apellido"]] is not None else ""
            patente = str(row[header_map["patente"]]).strip() if row[header_map["patente"]] is not None else ""
            comentario = str(row[header_map["comentario"]]).strip() if row[header_map["comentario"]] is not None else ""

            empresa = sector.nombre if sector else None

            # Saltar filas completamente vacías
            if not any([tipo_documento, rut, dni_extranjero, nombre, apellido, patente, comentario]):
                continue

            total += 1

            tipo_documento_normalizado = tipo_documento.replace(" ", "").upper()

            if tipo_documento_normalizado not in ["RUT", "DNI"]:
                errores.append({
                    "fila": idx,
                    "error": "TIPO DOCUMENTO inválido. Use RUT o DNI"
                })
                continue

            if not nombre:
                errores.append({"fila": idx, "error": "NOMBRE vacío"})
                continue

            if not apellido:
                errores.append({"fila": idx, "error": "APELLIDO vacío"})
                continue

            es_extranjero = tipo_documento_normalizado == "DNI"

            if es_extranjero:
                if not dni_extranjero:
                    errores.append({
                        "fila": idx,
                        "error": "DNI vacío para registro tipo DNI"
                    })
                    continue

                if Visita.objects.filter(
                        dni_extranjero=dni_extranjero,
                        es_extranjero=True
                ).exists():
                    errores.append({
                        "fila": idx,
                        "error": f"DNI duplicado: {dni_extranjero}"
                    })
                    continue

                rut = None

            else:
                if not rut:
                    errores.append({
                        "fila": idx,
                        "error": "RUT vacío para registro tipo RUT"
                    })
                    continue

                if Visita.objects.filter(
                        rut=rut,
                        es_extranjero=False
                ).exists():
                    errores.append({
                        "fila": idx,
                        "error": f"RUT duplicado: {rut}"
                    })
                    continue

                dni_extranjero = None

            try:
                Visita.objects.create(
                    rut=rut,
                    dni_extranjero=dni_extranjero,
                    es_extranjero=es_extranjero,
                    nombre=nombre,
                    apellido=apellido,
                    empresa=empresa,
                    patente=patente or None,
                    comentario=comentario or None,
                    sector=sector,
                    instalacion=instalacion,
                )
                creados += 1

            except IntegrityError:
                errores.append({
                    "fila": idx,
                    "error": "Error de integridad al guardar el registro"
                })
            except Exception as e:
                errores.append({
                    "fila": idx,
                    "error": str(e)
                })

        return Response({
            "total_filas_procesadas": total,
            "creados": creados,
            "errores": len(errores),
            "detalle_errores": errores,
            "sector_id": sector.id if sector else None,
        }, status=status.HTTP_200_OK)


class EnroladoDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        user = request.user

        try:
            visita = Visita.objects.get(id=pk)
        except Visita.DoesNotExist:
            return Response(
                {"detail": "Persona enrolada no encontrada"},
                status=status.HTTP_404_NOT_FOUND
            )

        # cliente sectorial: solo su sector
        if user.solo_enrolamiento:
            if visita.sector_id != user.sector_id:
                return Response(
                    {"detail": "No tiene permisos para eliminar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )

        # admin normal: solo su instalación
        elif not es_admin_general(user):
            if visita.instalacion_id != user.instalacion_id:
                return Response(
                    {"detail": "No tiene permisos para eliminar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )

        visita.delete()

        return Response(
            {"detail": "Registro eliminado correctamente"},
            status=status.HTTP_200_OK
        )


class EnroladoDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        user = request.user

        try:
            visita = Visita.objects.get(id=pk)
        except Visita.DoesNotExist:
            return Response(
                {"detail": "Persona enrolada no encontrada"},
                status=status.HTTP_404_NOT_FOUND
            )

        if user.solo_enrolamiento:
            if visita.sector_id != user.sector_id:
                return Response(
                    {"detail": "No tiene permisos para eliminar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )

        elif not es_admin_general(user):
            if visita.instalacion_id != user.instalacion_id:
                return Response(
                    {"detail": "No tiene permisos para eliminar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )

        visita.delete()

        return Response(
            {"detail": "Registro eliminado correctamente"},
            status=status.HTTP_200_OK
        )


class ProhibirAccesoEnroladoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        user = request.user

        try:
            visita = Visita.objects.get(id=pk)
        except Visita.DoesNotExist:
            return Response(
                {"detail": "Persona enrolada no encontrada"},
                status=status.HTTP_404_NOT_FOUND
            )

        if user.solo_enrolamiento:
            if visita.sector_id != user.sector_id:
                return Response(
                    {"detail": "No tiene permisos para prohibir el acceso de este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )
            instalacion = user.instalacion

        elif not es_admin_general(user):
            if visita.instalacion_id != user.instalacion_id:
                return Response(
                    {"detail": "No tiene permisos para prohibir el acceso de este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )
            instalacion = user.instalacion

        else:
            instalacion = visita.instalacion

        if not instalacion:
            return Response(
                {"detail": "No se pudo determinar la instalación para registrar la prohibición"},
                status=status.HTTP_400_BAD_REQUEST
            )

        prohibicion_activa = ProhibicionAcceso.objects.filter(
            visita=visita,
            instalacion=instalacion,
            fecha_fin__isnull=True
        ).exists()

        if prohibicion_activa:
            return Response(
                {"detail": "La persona ya tiene una prohibición activa en esta instalación"},
                status=status.HTTP_400_BAD_REQUEST
            )

        motivo = request.data.get("motivo", "").strip() or "Prohibición registrada desde módulo de enrolamiento"

        ProhibicionAcceso.objects.create(
            visita=visita,
            instalacion=instalacion,
            motivo=motivo,
            fecha_inicio=timezone.now(),
        )

        visita.estado = "prohibido"
        visita.save(update_fields=["estado", "actualizado_en"])

        return Response(
            {"detail": "Prohibición de acceso registrada correctamente"},
            status=status.HTTP_201_CREATED
        )


class DescargarPlantillaEnrolamientoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Enrolamiento"

        headers = [
            "TIPO DOCUMENTO",
            "RUT",
            "DNI",
            "NOMBRE",
            "APELLIDO",
            "PATENTE",
            "COMENTARIO",
        ]

        # Encabezados
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F2A24")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Anchos
        widths = [20, 18, 18, 24, 24, 16, 30]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Dropdown TIPO DOCUMENTO
        dv = DataValidation(type="list", formula1='"RUT,DNI"', allow_blank=False)
        dv.prompt = "Seleccione RUT para chilenos o DNI para extranjeros"
        dv.promptTitle = "Tipo de documento"
        dv.error = "Solo puede seleccionar RUT o DNI"
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)

        # Aplicar dropdown y valor por defecto
        for row in range(2, 301):
            cell_ref = f"A{row}"
            dv.add(cell_ref)
            ws[cell_ref] = "RUT"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_enrolamiento.xlsx"'

        wb.save(response)
        return response


class HabilitarAccesoEnroladoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        user = request.user

        try:
            visita = Visita.objects.get(id=pk)
        except Visita.DoesNotExist:
            return Response(
                {"detail": "Persona enrolada no encontrada"},
                status=status.HTTP_404_NOT_FOUND
            )

        if user.solo_enrolamiento:
            if visita.sector_id != user.sector_id:
                return Response(
                    {"detail": "No tiene permisos para habilitar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )
            instalacion = user.instalacion

        elif not es_admin_general(user):
            if visita.instalacion_id != user.instalacion_id:
                return Response(
                    {"detail": "No tiene permisos para habilitar este registro"},
                    status=status.HTTP_403_FORBIDDEN
                )
            instalacion = user.instalacion

        else:
            instalacion = visita.instalacion

        if not instalacion:
            return Response(
                {"detail": "No se pudo determinar la instalación"},
                status=status.HTTP_400_BAD_REQUEST
            )

        prohibiciones_activas = ProhibicionAcceso.objects.filter(
            visita=visita,
            instalacion=instalacion,
            fecha_fin__isnull=True
        )

        if not prohibiciones_activas.exists():
            return Response(
                {"detail": "La persona no tiene una prohibición activa en esta instalación"},
                status=status.HTTP_400_BAD_REQUEST
            )

        prohibiciones_activas.update(fecha_fin=timezone.now())

        visita.estado = "activo"
        visita.save(update_fields=["estado", "actualizado_en"])

        return Response(
            {"detail": "Restricción levantada correctamente"},
            status=status.HTTP_200_OK
        )
